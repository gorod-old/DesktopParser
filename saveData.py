import json
import os
import re
import shutil
from datetime import datetime
from PIL import Image
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, fills
from openpyxl.utils import get_column_letter
from striprtf.striprtf import rtf_to_text
from win32com.client import constants
from win32com.client.dynamic import CDispatch

from MessagePack import print_exception_msg, print_info_msg


def remove_directory(path: str):
    try:
        shutil.rmtree(path)
    except OSError as e:
        print_exception_msg(location='remove_directory', msg=f'{e.filename} - {e.strerror}')


def save_xlsx(p_data, file_name: str = 'result', folder: str = None,
              header_map=None, start_time: datetime = None,
              header_height: int = 30, row_height: int = 20, column_width: int = 15,
              col_width_map=None, auto_start: bool = False):
    """Saving data to xlsx file.
    Coloring the background of cells by key in cell values
    '[color_beige]': beige;
    '[color_red]': red;
    '[color_salad]': salad;
    '[color_green]': green;
    '[color_orange]': orange;
    '[color_blue]': blue;
    Hyperlink text in cell: [hyperlink]"""
    # hyper = ['№ аукциона', 'Ссылка на выписку из реестра МинПромТорга', 'Ссылка на РУ', 'Файлы']
    if len(p_data) == 0:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист 1"
    n = 3  # start row number without header (with a report from 1)
    h_row = n - 1
    row = n
    headers = {}
    # set header height
    ws.row_dimensions[h_row].height = header_height
    # set header borders
    h_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    if header_map:
        i = 1
        for key in header_map:
            headers.update({key: i})
            ws.cell(row=h_row, column=len(headers)).value = key
            ws.cell(row=h_row, column=len(headers)).border = h_border
            i += 1
    for i, data in enumerate(p_data):
        color = __get_xlsx_fill('[color_beige]') if i % 2 == 0 else __get_xlsx_fill('[color_lgreen]')
        for data_row in data:
            for j in range(len(headers)):
                ws.cell(row=row, column=j + 1).fill = color
            for key in data_row.keys():
                val = data_row[key]
                # val = (val, '[hyperlink]') if key in hyper else (val,)
                col = headers[key]
                __set_xlsx_cell(ws, row=row, column=col, value=val)
            ws.row_dimensions[row].height = row_height
            row += 1
    # columns width and borders
    for key in headers.keys():
        ws.column_dimensions[get_column_letter(headers[key])].width = column_width \
            if col_width_map is None or col_width_map.get(key) is None else col_width_map.get(key)
        ws.cell(row=row - 1, column=headers[key]).border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
        for j in range(n, row - 1):
            ws.cell(row=j, column=headers[key]).border = Border(right=Side(style='thin'))
    if start_time:
        start_time = start_time.strftime("%d/%m/%Y %H:%M:%S"). \
            replace('/', '-').replace(' ', '_').replace(':', '-')
        end_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S"). \
            replace('/', '-').replace(' ', '_').replace(':', '-')
        ws.cell(row=1, column=1).value = f'Начало парсинга: {start_time}, конец парсинга: {end_time}'
        ws.cell(row=1, column=1).border = Border()
        # colorize file info
        for i in range(len(headers)):
            ws.cell(row=1, column=i + 1).fill = __get_xlsx_fill('[color_beige]')

    # save file
    path = __get_file_path('xlsx', file_name=file_name, folder=folder)
    print_info_msg(location='save_xlsx', msg=f'save path: {path}')
    if os.path.exists(path):
        os.remove(path)
    wb.save(path)
    if auto_start:
        os.startfile(path)


def __set_xlsx_cell(ws, row: int, column: int, value):
    if type(value) != tuple and type(value) != list:
        value = ((value,),)
    # beige color (even line)
    fill_beige = PatternFill(start_color='f5eeda',
                             end_color='f5eeda',
                             fill_type='solid')
    # red color
    fill_red = PatternFill(start_color='f5c7bf',
                           end_color='f5c7bf',
                           fill_type='solid')
    # salad color
    fill_salad = PatternFill(start_color='e8ffdb',
                             end_color='e8ffdb',
                             fill_type='solid')
    # green
    fill_green = PatternFill(start_color='a4f59d',
                             end_color='a4f59d',
                             fill_type='solid')
    # light green
    fill_lgreen = PatternFill(start_color='c3fad2',
                              end_color='c3fad2',
                              fill_type='solid')
    # orange
    fill_orange = PatternFill(start_color='fce079',
                              end_color='fce079',
                              fill_type='solid')
    # blue
    fill_blue = PatternFill(start_color='cfe5fa',
                            end_color='cfe5fa',
                            fill_type='solid')
    # white
    fill_white = PatternFill(start_color='ffffff',
                             end_color='ffffff',
                             fill_type='solid')
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    args = {
        '[color_beige]': fill_beige,
        '[color_red]': fill_red,
        '[color_salad]': fill_salad,
        '[color_green]': fill_green,
        '[color_lgreen]': fill_lgreen,
        '[color_orange]': fill_orange,
        '[color_blue]': fill_blue,
        '[color_white]': fill_white,
        '[hyperlink]': value[0][1] if len(value[0]) > 1 else value[0][0],
    }
    ws.cell(row=row, column=column).value = value[0][0]
    ws.cell(row=row, column=column).border = border
    for i in range(1, len(value)):
        arg = args.get(value[i])
        if arg and type(arg) == fills.PatternFill:
            ws.cell(row=row, column=column).fill = arg
        elif value[i] == '[hyperlink]':
            ws.cell(row=row, column=column).hyperlink = arg


def __get_xlsx_fill(key):
    # beige color (even line)
    fill_beige = PatternFill(start_color='f5eeda',
                             end_color='f5eeda',
                             fill_type='solid')
    # red color
    fill_red = PatternFill(start_color='f5c7bf',
                           end_color='f5c7bf',
                           fill_type='solid')
    # salad color
    fill_salad = PatternFill(start_color='e8ffdb',
                             end_color='e8ffdb',
                             fill_type='solid')
    # green
    fill_green = PatternFill(start_color='a4f59d',
                             end_color='a4f59d',
                             fill_type='solid')
    # light green
    fill_lgreen = PatternFill(start_color='c3fad2',
                              end_color='c3fad2',
                              fill_type='solid')
    # orange
    fill_orange = PatternFill(start_color='fce079',
                              end_color='fce079',
                              fill_type='solid')
    # blue
    fill_blue = PatternFill(start_color='cfe5fa',
                            end_color='cfe5fa',
                            fill_type='solid')
    # white
    fill_white = PatternFill(start_color='ffffff',
                             end_color='ffffff',
                             fill_type='solid')
    fill_list = {
        '[color_beige]': fill_beige,
        '[color_red]': fill_red,
        '[color_salad]': fill_salad,
        '[color_green]': fill_green,
        '[color_lgreen]': fill_lgreen,
        '[color_orange]': fill_orange,
        '[color_blue]': fill_blue,
        '[color_white]': fill_white,
    }
    return fill_list.get(key) or PatternFill()


def save_json(json_data, root_folder: str = '', file_name: str = 'result', encoding='utf-8', folder: str = None):
    path = __get_file_path('json', root_folder=root_folder, file_name=file_name, folder=folder)
    print_info_msg(location='save_json', msg=f'save path: {path}')
    if os.path.exists(path):
        os.remove(path)
    with open(path, 'a', encoding=encoding) as file:
        json.dump(json_data, file, indent=4, ensure_ascii=False)


def get_json_data_from_file(path, encoding='utf-8', stream: int = None):
    print_info_msg(location='get_json_data_from_file', msg=f'get data path: {path}', stream=stream)
    json_content = open(path, 'r', encoding=encoding).read()
    json_data = json.loads(json_content)
    return json_data


def __get_file_path(extension: str, root_folder: str = 'result data', file_name: str = 'result', folder: str = None):
    if root_folder is None or root_folder == '':
        print_info_msg(location='__get_file_path', msg=f'no root folder specified for file: {file_name}.{extension}. '
                                                       f'Set to: None')
        root_folder = None
    root = os.getcwd() + f'/{root_folder}' if root_folder is not None else os.getcwd()
    if not os.path.exists(root) or not os.path.isdir(root):
        os.mkdir(root)
    folder = '/' + folder if folder is not None else ''
    if folder != '' and (not os.path.exists(root + folder) or not os.path.isdir(root + folder)):
        os.mkdir(root + folder)
    path = root + folder + f'/{file_name}.{extension}'
    return os.path.normpath(path)


def get_text_from_file(path, encoding='utf-8', stream: int = None):
    print_info_msg(location='get_text_from_file', msg=f'get data path: {path}', stream=stream)
    with open(path, 'r', encoding=encoding) as file:
        return file.read()


def rtf_to_text_(rtf, stream: int = None):
    print_info_msg(location='rtf_to_text', msg=f'convert rtf to text', stream=stream)
    return rtf_to_text(rtf, 'ignore')


def __get_file_name(path: str):
    return path.replace(os.sep, '/').split('/')[-1].split('.')[0].replace(' ', '_').lower()


def get_office_app_instance(app_id=None) -> CDispatch:
    # Initialize
    win32.pythoncom.CoInitialize()
    # Get instance from the id
    app = None
    if app_id:
        app = win32.Dispatch(
            win32.pythoncom.CoGetInterfaceAndReleaseStream(app_id, win32.pythoncom.IID_IDispatch)
        )
    return app


def save_as_docx(path, word_instance: CDispatch = None, stream: int = None):
    # Opening MS Word
    word = win32.gencache.EnsureDispatch('Word.Application') if word_instance is None else word_instance
    doc = word.Documents.Open(path)
    doc.Activate()

    # Rename path with .docx
    new_file_abs = os.path.abspath(path)
    new_file_abs = re.sub(r'\.\w+$', '.docx', new_file_abs)

    # Save and Close
    word.ActiveDocument.SaveAs(
        new_file_abs, FileFormat=constants.wdFormatXMLDocument
    )
    doc.Close(False)
    print_info_msg(location='save_as_docx', msg=f'docx file path: {new_file_abs}', stream=stream)
    return new_file_abs


def resize_image(path: str, max_size: int):
    if max_size is None:
        return
    # open image file with PIL
    img = Image.open(path)
    if img.size[1] >= img.size[0] and img.size[1] > max_size:
        # by max height
        w_size = int(float(img.size[0] * max_size) / float(img.size[1]))
        img = img.resize((w_size, max_size), Image.ANTIALIAS)
    elif img.size[0] > img.size[1] and img.size[0] > max_size:
        # by max width
        h_size = int(float(img.size[1] * max_size) / float(img.size[0]))
        img = img.resize((max_size, h_size), Image.ANTIALIAS)
    img.save(path)
